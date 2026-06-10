"""
YouTube チャンネルスクリーニングの中核ロジック（CLI / Web 共有モジュール）。

検索 → チャンネル詳細 → フィルタ → テーマ判定 → スコア → 出力 という流れを提供する。
CLI(main.py) と FastAPI バックエンド(app.py) の両方からこのモジュールを使う。

クオータ節約のため、検索結果・チャンネル詳細・動画統計をローカルにキャッシュする。
"""
from __future__ import annotations

import math
import time
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple

import json

import httplib2
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

# OS の証明書ストアで TLS 検証する（社内プロキシ/セキュリティソフトが証明書を
# 差し込む環境では、これが無いと certifi/httplib2 同梱CAでは検証に失敗する）。
try:
    import truststore

    truststore.inject_into_ssl()
except Exception:  # noqa: BLE001  未インストール等でも致命的にしない
    pass
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# テーマ判定用キーワード（部分一致）。matched キーワードや ch 名/説明から推定。
THEME_FISHING = ["釣り", "リール", "魚探", "魚群探知", "船", "沖", "ジギング",
                 "タイラバ", "ボート", "深場", "中深場", "ホンデックス", "ガーミン",
                 "GPS魚探", "ジグ", "アングラー", "フィッシング"]
THEME_ASTRO = ["星", "天体", "赤道儀", "電視観望", "星雲", "天の川", "ディープスカイ",
               "望遠鏡", "星景", "星空", "観望", "天文", "ポタ赤", "タイムラプス"]

# この秒数以下の動画はショート扱い（YouTube APIに公式フラグは無く、長さで判定する）。
# ショートは現在最長3分なので 180 秒。1〜3分の通常動画をショート扱いする誤分類は許容。
SHORT_MAX_SECONDS = 180

NoOp = lambda *a, **k: None  # noqa: E731


# --------------------------------------------------------------------------- #
# キャッシュ
# --------------------------------------------------------------------------- #
class Cache:
    """単純な JSON ファイルキャッシュ（種類ごとに 1 ファイル、TTL 付き）。"""

    def __init__(self, directory: Path, ttl_seconds: int) -> None:
        self.directory = Path(directory)
        self.ttl = ttl_seconds
        self.directory.mkdir(parents=True, exist_ok=True)
        self._stores: Dict[str, Dict[str, Any]] = {}

    def _path(self, store: str) -> Path:
        return self.directory / f"{store}.json"

    def _load(self, store: str) -> Dict[str, Any]:
        if store not in self._stores:
            path = self._path(store)
            if path.exists():
                try:
                    self._stores[store] = json.loads(path.read_text(encoding="utf-8"))
                except (json.JSONDecodeError, OSError):
                    self._stores[store] = {}
            else:
                self._stores[store] = {}
        return self._stores[store]

    def get(self, store: str, key: str) -> Optional[Any]:
        entry = self._load(store).get(key)
        if not entry:
            return None
        if time.time() - entry.get("_ts", 0) > self.ttl:
            return None  # 期限切れ
        return entry.get("value")

    def _write(self, store: str) -> None:
        try:
            self._path(store).write_text(
                json.dumps(self._stores[store], ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        except OSError:
            # 書き込み不可な環境（読み取り専用FS等）でもインメモリでは効かせる
            pass

    def set(self, store: str, key: str, value: Any) -> None:
        self._load(store)[key] = {"_ts": time.time(), "value": value}
        self._write(store)

    def set_many(self, store: str, items: Dict[str, Any]) -> None:
        """複数件をまとめて保存（ファイル書き込みは1回）。"""
        if not items:
            return
        data = self._load(store)
        now = time.time()
        for k, v in items.items():
            data[k] = {"_ts": now, "value": v}
        self._write(store)


# --------------------------------------------------------------------------- #
# YouTube API ラッパー
# --------------------------------------------------------------------------- #
class YouTubeScreener:
    def __init__(self, api_key: str, cache: Cache, log: Callable[[str], None] = NoOp) -> None:
        self.youtube = build("youtube", "v3", developerKey=api_key, cache_discovery=False)
        self.cache = cache
        self.log = log
        self.quota_used = 0       # 概算クオータ消費（キャッシュヒット分は加算しない）
        self.api_searches = 0     # 実際に API を叩いた検索回数

    def search_channels(
        self, keyword: str, region: str, language: str, order: str,
        max_results: int, search_budget_left: int,
    ) -> Tuple[List[str], bool]:
        """キーワードでチャンネルを検索し (channelId のリスト, API を叩いたか) を返す。"""
        cache_key = f"{region}|{language}|{order}|{max_results}|{keyword}"
        cached = self.cache.get("search", cache_key)
        if cached is not None:
            self.log(f"[cache] 検索: {keyword!r}")
            return cached, False

        if search_budget_left <= 0:
            self.log(f"[skip] 検索上限のためスキップ: {keyword!r}")
            return [], False

        channel_ids: List[str] = []
        page_token: Optional[str] = None
        while len(channel_ids) < max_results:
            resp = (
                self.youtube.search()
                .list(
                    q=keyword,
                    part="snippet",
                    type="channel",
                    regionCode=region,
                    relevanceLanguage=language,
                    order=order,
                    maxResults=min(50, max_results - len(channel_ids)),
                    pageToken=page_token,
                )
                .execute()
            )
            self.quota_used += 100  # search.list は 100 ユニット
            for item in resp.get("items", []):
                cid = item["snippet"]["channelId"]
                if cid not in channel_ids:
                    channel_ids.append(cid)
            page_token = resp.get("nextPageToken")
            if not page_token:
                break

        self.api_searches += 1
        self.cache.set("search", cache_key, channel_ids)
        self.log(f"[API] 検索: {keyword!r} -> {len(channel_ids)} 件")
        return channel_ids, True

    def get_channel_details(self, channel_ids: List[str]) -> Dict[str, Dict[str, Any]]:
        """channelId -> 詳細 dict。最大50件ずつ channels.list で取得。"""
        result: Dict[str, Dict[str, Any]] = {}
        to_fetch: List[str] = []
        for cid in channel_ids:
            cached = self.cache.get("channels", cid)
            if cached is not None:
                result[cid] = cached
            else:
                to_fetch.append(cid)

        for batch in _chunked(to_fetch, 50):
            resp = (
                self.youtube.channels()
                .list(part="snippet,statistics,contentDetails", id=",".join(batch))
                .execute()
            )
            self.quota_used += 1  # channels.list は 1 ユニット
            new_items: Dict[str, Any] = {}
            for item in resp.get("items", []):
                stats = item.get("statistics", {})
                details = {
                    "channel_id": item["id"],
                    "title": item["snippet"]["title"],
                    "description": item["snippet"].get("description", ""),
                    "country": item["snippet"].get("country", ""),
                    "published_at": item["snippet"].get("publishedAt"),
                    "subscriber_count": _to_int(stats.get("subscriberCount")),
                    "subscriber_hidden": stats.get("hiddenSubscriberCount", False),
                    "view_count": _to_int(stats.get("viewCount")),
                    "video_count": _to_int(stats.get("videoCount")),
                    "uploads_playlist": item["contentDetails"]["relatedPlaylists"].get("uploads"),
                    "url": f"https://www.youtube.com/channel/{item['id']}",
                }
                result[item["id"]] = details
                new_items[item["id"]] = details
            self.cache.set_many("channels", new_items)  # バッチごとに1回だけ書き込み
        return result

    def get_recent_stats_batch(
        self, channels: List[Dict[str, Any]], count: int,
        progress: Callable[[int, int], None] = NoOp, workers: int = 8,
    ) -> Dict[str, Dict[str, Optional[Any]]]:
        """各チャンネルの直近 count 本の {平均再生数, 最終投稿日} をまとめて取得する。

        最適化:
        - 再生リスト(playlistItems)取得はチャンネル単位で必須だが **並列**で実行
        - 動画統計(videos.list)は全チャンネル横断で **50件ずつ一括**取得し、動画単位でキャッシュ
        - 既に recent キャッシュがあるチャンネルは API を一切叩かない
        """
        empty = {
            "avg_views": None, "avg_views_long": None, "avg_views_short": None,
            "long_count": 0, "short_count": 0, "last_upload": None, "emails": [],
        }
        results: Dict[str, Dict[str, Optional[Any]]] = {}
        cid_to_pl: Dict[str, Optional[str]] = {}
        pending: List[Tuple[str, str]] = []  # (channel_id, uploads_playlist)

        # recent キャッシュキーに本数とショート閾値を含める＝設定を変えたら自動で作り直される
        rkey = lambda pl: f"{pl}|{count}|v2s{SHORT_MAX_SECONDS}"  # noqa: E731

        for ch in channels:
            cid = ch["channel_id"]
            pl = ch.get("uploads_playlist")
            cid_to_pl[cid] = pl
            if not pl:
                results[cid] = dict(empty)
                continue
            cached = self.cache.get("recent", rkey(pl))
            if cached is not None:
                results[cid] = cached
            else:
                pending.append((cid, pl))

        total = len(pending)
        progress(0, total)

        # 1) playlistItems を並列取得（各スレッドは専用 Http を使う：httplib2 は非スレッドセーフ）
        def fetch(item: Tuple[str, str]):
            cid, pl = item
            req = self.youtube.playlistItems().list(
                part="contentDetails", playlistId=pl, maxResults=count
            )
            try:
                resp = req.execute(http=httplib2.Http(timeout=30))
            except HttpError:
                return cid, None, None
            items = resp.get("items", [])
            vids = [it["contentDetails"]["videoId"] for it in items]
            dates = [it["contentDetails"].get("videoPublishedAt") for it in items]
            return cid, vids, max((d for d in dates if d), default=None)

        channel_videos: Dict[str, Tuple[Optional[List[str]], Optional[str]]] = {}
        if pending:
            done = 0
            with ThreadPoolExecutor(max_workers=min(workers, len(pending))) as ex:
                for cid, vids, last in ex.map(fetch, pending):
                    self.quota_used += 1  # playlistItems.list は 1 ユニット（メインスレッドで加算）
                    channel_videos[cid] = (vids, last)
                    done += 1
                    progress(done, total)

        # 2) 全動画IDを集約し、動画単位キャッシュを除いて 50件ずつ一括取得
        ordered_ids: List[str] = []
        seen = set()
        for vids, _ in channel_videos.values():
            for v in vids or []:
                if v not in seen:
                    seen.add(v)
                    ordered_ids.append(v)

        # 動画キャッシュは {"views","emails","dur"}。dur 欠落（旧形式/旧版）は取り直す。
        view_map: Dict[str, Optional[int]] = {}
        email_map: Dict[str, List[str]] = {}
        dur_map: Dict[str, Optional[int]] = {}
        to_fetch: List[str] = []
        for vid in ordered_ids:
            c = self.cache.get("videos", vid)
            if isinstance(c, dict) and "dur" in c:
                view_map[vid] = c.get("views")
                email_map[vid] = c.get("emails", [])
                dur_map[vid] = c.get("dur")
            else:
                to_fetch.append(vid)  # 旧int / dur無し dict は再取得

        new_cache: Dict[str, Any] = {}
        for batch in _chunked(to_fetch, 50):
            # snippet/contentDetails を足してもクオータは1のまま（概要欄メール＋動画の長さ）。
            resp = (
                self.youtube.videos()
                .list(part="statistics,snippet,contentDetails", id=",".join(batch))
                .execute()
            )
            self.quota_used += 1  # videos.list は 1 ユニット（50件まとめて）
            for it in resp.get("items", []):
                v = _to_int(it.get("statistics", {}).get("viewCount"))
                emails = extract_emails(it.get("snippet", {}).get("description", ""))
                dur = _duration_seconds(it.get("contentDetails", {}).get("duration"))
                view_map[it["id"]] = v
                email_map[it["id"]] = emails
                dur_map[it["id"]] = dur
                new_cache[it["id"]] = {"views": v, "emails": emails, "dur": dur}
        self.cache.set_many("videos", new_cache)

        # 3) チャンネルごとに長尺/ショートを仕分けて平均、recent キャッシュへ一括保存
        recent_to_cache: Dict[str, Any] = {}
        for cid, (vids, last) in channel_videos.items():
            if vids is None:  # playlistItems が失敗
                results[cid] = dict(empty)
                continue
            long_views: List[int] = []
            short_views: List[int] = []
            for v in vids:
                views = view_map.get(v)
                if views is None:
                    continue
                dur = dur_map.get(v)
                if dur is not None and dur <= SHORT_MAX_SECONDS:
                    short_views.append(views)
                else:  # 長尺、または長さ不明は長尺扱い
                    long_views.append(views)
            avg_long = round(sum(long_views) / len(long_views), 1) if long_views else None
            avg_short = round(sum(short_views) / len(short_views), 1) if short_views else None
            emails: List[str] = []
            for v in vids:
                for e in email_map.get(v, []):
                    if e not in emails:
                        emails.append(e)
            out = {
                "avg_views": avg_long,        # フィルタ/スコアの基準＝長尺平均
                "avg_views_long": avg_long,
                "avg_views_short": avg_short,
                "long_count": len(long_views),
                "short_count": len(short_views),
                "last_upload": last,
                "emails": emails,
            }
            results[cid] = out
            pl = cid_to_pl.get(cid)
            if pl:
                recent_to_cache[rkey(pl)] = out
        self.cache.set_many("recent", recent_to_cache)

        return results


# --------------------------------------------------------------------------- #
# ヘルパー
# --------------------------------------------------------------------------- #
def _chunked(items: List[str], size: int) -> Iterable[List[str]]:
    for i in range(0, len(items), size):
        yield items[i : i + size]


def _to_int(value: Any) -> Optional[int]:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


import re

_DUR_RE = re.compile(r"PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?$")


def _duration_seconds(iso: Optional[str]) -> Optional[int]:
    """ISO8601 duration（例 PT2M30S）を秒に変換。判定不能なら None。"""
    if not iso:
        return None
    m = _DUR_RE.match(iso)
    if not m:
        return None
    h, mi, s = (int(x) if x else 0 for x in m.groups())
    return h * 3600 + mi * 60 + s

_EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
# 画像/動画ファイル名などの誤検出を弾く
_EMAIL_NG_SUFFIX = (".png", ".jpg", ".jpeg", ".gif", ".webp", ".mp4", ".com.")


def extract_emails(text: Optional[str]) -> List[str]:
    """テキスト（チャンネル説明/動画概要）からメールアドレスを抽出（重複除去）。"""
    if not text:
        return []
    found = []
    for m in _EMAIL_RE.findall(text):
        e = m.strip().lower()
        if e.endswith(_EMAIL_NG_SUFFIX):
            continue
        if e not in found:
            found.append(e)
    return found


def _parse_dt(iso: Optional[str]) -> Optional[datetime]:
    if not iso:
        return None
    try:
        return datetime.fromisoformat(iso.replace("Z", "+00:00"))
    except ValueError:
        return None


def _format_date(iso: Optional[str]) -> str:
    dt = _parse_dt(iso)
    return dt.astimezone(timezone.utc).strftime("%Y-%m-%d") if dt else (iso or "")


def _days_since(iso: Optional[str]) -> Optional[int]:
    dt = _parse_dt(iso)
    if not dt:
        return None
    return (datetime.now(timezone.utc) - dt.astimezone(timezone.utc)).days


def _classify_theme(text: str) -> str:
    fishing = sum(1 for kw in THEME_FISHING if kw in text)
    astro = sum(1 for kw in THEME_ASTRO if kw in text)
    if fishing == 0 and astro == 0:
        return "不明"
    if fishing > astro:
        return "釣り"
    if astro > fishing:
        return "星空"
    return "釣り/星空"


# --------------------------------------------------------------------------- #
# 設定
# --------------------------------------------------------------------------- #
class Config:
    """dict（YAML or JSON リクエスト）から設定を読む。"""

    def __init__(self, data: Dict[str, Any]) -> None:
        data = data or {}
        self.region = data.get("region_code", "JP")
        self.language = data.get("language", "ja")
        self.results_per_keyword = int(data.get("results_per_keyword", 25))
        self.search_order = data.get("search_order", "relevance")
        self.recent_video_count = int(data.get("recent_video_count", 10))

        # 階層化キーワード（無ければ従来の flat keywords も許容）
        self.keyword_tier: Dict[str, int] = {}
        for tier, field in [(1, "keywords_tier1"), (2, "keywords_tier2"), (3, "keywords_tier3")]:
            for kw in data.get(field, []) or []:
                kw = (kw or "").strip()
                if kw:
                    self.keyword_tier.setdefault(kw, tier)
        for kw in data.get("keywords", []) or []:   # 後方互換
            kw = (kw or "").strip()
            if kw:
                self.keyword_tier.setdefault(kw, 1)
        self.keywords = list(self.keyword_tier.keys())

        self.exclude_title_keywords = data.get("exclude_title_keywords", []) or []
        self.competitor_flag_keywords = data.get("competitor_flag_keywords", []) or []

        f = data.get("filters", {}) or {}
        self.sub_min = f.get("subscriber_min")
        self.sub_max = f.get("subscriber_max")
        self.avg_views_min = f.get("avg_views_min")
        self.engagement_min = f.get("engagement_min")
        self.min_video_count = f.get("min_video_count")
        self.active_within_days = f.get("active_within_days")

        self.theme_overrides = data.get("theme_overrides", {}) or {}
        self.scoring_weights = data.get("scoring_weights", {}) or {}
        self.multi_keyword_bonus = bool(data.get("multi_keyword_bonus", False))

        self.cache_ttl_hours = int(data.get("cache_ttl_hours", 168))
        self.max_searches_per_run = int(data.get("max_searches_per_run", 100))
        self.output_file = data.get("output_file", "output.xlsx")


# --------------------------------------------------------------------------- #
# スコアリング
# --------------------------------------------------------------------------- #
def compute_score(info: Dict[str, Any], cfg: Config) -> float:
    w = cfg.scoring_weights
    tiers = {cfg.keyword_tier.get(kw, 3) for kw in info["matched_keywords"]}
    best_tier = min(tiers) if tiers else 3
    scene = {1: 1.0, 2: 0.6, 3: 0.3}.get(best_tier, 0.0)

    eng = info.get("engagement") or 0.0
    eng_norm = min(eng / 0.15, 1.0)

    days = info.get("days_since_upload")
    window = cfg.active_within_days or 180
    act_norm = max(0.0, 1.0 - days / window) if days is not None else 0.0

    subs = info.get("subscriber_count") or 0
    fit_norm = 0.0
    if subs > 0 and cfg.sub_min and cfg.sub_max and cfg.sub_max > cfg.sub_min:
        lo, hi = math.log(cfg.sub_min), math.log(cfg.sub_max)
        mid = (lo + hi) / 2
        fit_norm = max(0.0, 1.0 - abs(math.log(subs) - mid) / ((hi - lo) / 2))

    score = (
        scene * w.get("scene_match", 0)
        + eng_norm * w.get("engagement", 0)
        + act_norm * w.get("recent_activity", 0)
        + fit_norm * w.get("subscriber_fit", 0)
    )
    if cfg.multi_keyword_bonus and 1 in tiers and (2 in tiers or 3 in tiers):
        score += 5
    return round(score, 1)


# --------------------------------------------------------------------------- #
# スクリーニング実行（CLI / Web 共通）
# --------------------------------------------------------------------------- #
ProgressCb = Callable[[str, int, int], None]  # (phase, done, total)


def run_screening(
    cfg: Config, screener: YouTubeScreener, progress: ProgressCb = NoOp,
) -> Dict[str, Any]:
    """設定に従ってスクリーニングを実行し {results, stats} を返す。"""
    # 1. 検索
    keyword_hits: Dict[str, List[str]] = {}
    total_kw = len(cfg.keywords)
    for i, kw in enumerate(cfg.keywords):
        budget = cfg.max_searches_per_run - screener.api_searches
        ids, _ = screener.search_channels(
            kw, cfg.region, cfg.language, cfg.search_order,
            cfg.results_per_keyword, budget,
        )
        for cid in ids:
            keyword_hits.setdefault(cid, []).append(kw)
        progress("searching", i + 1, total_kw)

    channel_ids = list(keyword_hits.keys())

    # 2. チャンネル詳細
    progress("details", 0, len(channel_ids))
    details = screener.get_channel_details(channel_ids)
    progress("details", len(channel_ids), len(channel_ids))

    # 3. 安いフィルタ（登録者数 / 動画本数 / タイトル除外語）
    drop = {"hidden": 0, "subs": 0, "videos": 0, "excluded": 0}
    stage1: List[Dict[str, Any]] = []
    for cid, info in details.items():
        if info.get("subscriber_hidden") or info.get("subscriber_count") is None:
            drop["hidden"] += 1
            continue
        subs = info["subscriber_count"]
        if cfg.sub_min is not None and subs < cfg.sub_min:
            drop["subs"] += 1
            continue
        if cfg.sub_max is not None and subs > cfg.sub_max:
            drop["subs"] += 1
            continue
        if cfg.min_video_count is not None and (info.get("video_count") or 0) < cfg.min_video_count:
            drop["videos"] += 1
            continue
        text = f"{info['title']} {info.get('description', '')}"
        if any(ng in text for ng in cfg.exclude_title_keywords):
            drop["excluded"] += 1
            continue
        info["matched_keywords"] = keyword_hits.get(cid, [])
        stage1.append(info)

    # 4. 直近動画を一括取得（playlistItems 並列 + videos.list 横断50件一括）
    recent_map = screener.get_recent_stats_batch(
        stage1, cfg.recent_video_count,
        progress=lambda done, total: progress("scoring", done, total),
    )

    # 5. 高いフィルタ → テーマ/競合/スコア（ここはAPIを叩かない）
    final: List[Dict[str, Any]] = []
    drop2 = {"avg": 0, "eng": 0, "active": 0, "relevance": 0}
    for info in stage1:
        recent = recent_map.get(info["channel_id"]) or {}
        avg = recent.get("avg_views")  # = 長尺平均
        info["avg_views"] = avg
        info["avg_views_short"] = recent.get("avg_views_short")
        info["long_count"] = recent.get("long_count", 0)
        info["short_count"] = recent.get("short_count", 0)
        info["last_upload"] = recent.get("last_upload")
        info["days_since_upload"] = _days_since(recent.get("last_upload"))
        subs = info.get("subscriber_count") or 0
        info["engagement"] = round(avg / subs, 4) if (avg and subs) else None

        # 連絡先メール: チャンネル説明＋動画概要から抽出（公開データのみ）
        emails: List[str] = extract_emails(info.get("description", ""))
        for e in recent.get("emails", []) or []:
            if e not in emails:
                emails.append(e)
        info["emails"] = emails
        info["email"] = emails[0] if emails else ""

        theme_text = " ".join(info["matched_keywords"]) + " " + info["title"] + " " + info.get("description", "")
        info["theme"] = _classify_theme(theme_text)
        override = cfg.theme_overrides.get(info["theme"], {}) or {}
        avg_min = override.get("avg_views_min", cfg.avg_views_min)
        eng_min = override.get("engagement_min", cfg.engagement_min)

        require = override.get("require_keywords")
        if require and not any(k in theme_text for k in require):
            drop2["relevance"] += 1
            continue
        if avg_min is not None and (avg is None or avg < avg_min):
            drop2["avg"] += 1
            continue
        if eng_min is not None and (info["engagement"] is None or info["engagement"] < eng_min):
            drop2["eng"] += 1
            continue
        if cfg.active_within_days is not None:
            d = info["days_since_upload"]
            if d is None or d > cfg.active_within_days:
                drop2["active"] += 1
                continue

        info["competitor_flags"] = [kw for kw in cfg.competitor_flag_keywords if kw in theme_text]
        info["score"] = compute_score(info, cfg)
        final.append(info)

    final.sort(key=lambda x: x.get("score") or 0, reverse=True)
    return {
        "results": final,
        "stats": {
            "unique_channels": len(channel_ids),
            "after_cheap_filter": len(stage1),
            "final": len(final),
            "drop_stage1": drop,
            "drop_stage2": drop2,
            "quota_used": screener.quota_used,
            "api_searches": screener.api_searches,
        },
    }


# --------------------------------------------------------------------------- #
# Excel 出力
# --------------------------------------------------------------------------- #
EXCEL_HEADERS = [
    ("総合スコア", "score", 10),
    ("チャンネル名", "title", 30),
    ("URL", "url", 45),
    ("登録者数", "subscriber_count", 12),
    ("長尺平均再生数", "avg_views", 14),
    ("ショート平均再生数", "avg_views_short", 16),
    ("長尺本数", "long_count", 9),
    ("ショート本数", "short_count", 11),
    ("エンゲージ率", "engagement", 12),
    ("動画本数", "video_count", 10),
    ("最終投稿日", "last_upload", 12),
    ("開設日", "published_at", 12),
    ("総再生数", "view_count", 14),
    ("メール候補(自動抽出)", "emails", 30),
    ("ヒットしたキーワード", "matched_keywords", 30),
    ("競合言及フラグ", "competitor_flags", 18),
    ("テーマ", "theme", 10),
    ("連絡先(手入力)", "_contact", 18),
    ("備考(手入力)", "_note", 24),
]


def _build_workbook(rows: List[Dict[str, Any]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "channels"

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    for col, (label, _key, width) in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width

    for r, row in enumerate(rows, start=2):
        for c, (_label, key, _width) in enumerate(EXCEL_HEADERS, start=1):
            value = row.get(key)
            if key in ("published_at", "last_upload"):
                value = _format_date(value)
            elif key in ("matched_keywords", "competitor_flags", "emails") and isinstance(value, list):
                value = ", ".join(value)
            cell = ws.cell(row=r, column=c, value=value)
            if key in ("subscriber_count", "view_count", "video_count", "long_count", "short_count"):
                cell.number_format = "#,##0"
            elif key in ("avg_views", "avg_views_short"):
                cell.number_format = "#,##0.0"
            elif key == "engagement":
                cell.number_format = "0.0%"

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_HEADERS))}{len(rows) + 1}"
    return wb


def write_excel(rows: List[Dict[str, Any]], output_path: Path, log: Callable[[str], None] = NoOp) -> Path:
    """Excel をファイルに保存し、実際に書き込んだパスを返す（CLI用）。

    出力先がロックされている場合は日時付きの別名に逃がす。
    """
    wb = _build_workbook(rows)
    try:
        wb.save(output_path)
        return output_path
    except PermissionError:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt_path = output_path.with_name(f"{output_path.stem}_{stamp}{output_path.suffix}")
        wb.save(alt_path)
        log(f"注意: {output_path.name} が開かれていて上書きできないため別名で保存しました。")
        return alt_path


def excel_bytes(rows: List[Dict[str, Any]]) -> bytes:
    """Excel をメモリ上で生成してバイト列で返す（Web ダウンロード用）。"""
    buf = BytesIO()
    _build_workbook(rows).save(buf)
    return buf.getvalue()
